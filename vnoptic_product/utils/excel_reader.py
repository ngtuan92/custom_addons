# -*- coding: utf-8 -*-
"""
Excel file reader utility
Parses Excel files and extracts product data
"""
import base64
import io
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def detect_product_type(sheet):
    """
    Detect product type from row 3, cell D3
    
    Args:
        sheet: openpyxl worksheet
    
    Returns:
        str: 'lens', 'opt', or 'accessory'
    
    Raises:
        ValueError: If product type cannot be detected
    """
    title_cell = sheet['D3'].value
    
    if not title_cell:
        raise ValueError("Cannot detect product type: Cell D3 is empty")
    
    title = str(title_cell).upper()
    
    if 'MẮT' in title or 'MAT' in title:
        return 'lens'
    elif 'GỌNG' in title or 'GONG' in title or 'KÍNH' in title or 'KINH' in title:
        return 'opt'
    elif 'PHỤ KIỆN' in title or 'PHU KIEN' in title:
        return 'accessory'
    else:
        raise ValueError(f"Unknown product type in cell D3: {title_cell}")


def parse_headers(sheet, header_row=10):
    """
    Parse English headers from specified row
    
    Args:
        sheet: openpyxl worksheet
        header_row: Row number containing English headers (default: 10)
    
    Returns:
        list: List of header names
    """
    headers = []
    col_idx = 1
    
    while True:
        cell_value = sheet.cell(row=header_row, column=col_idx).value
        if cell_value is None or str(cell_value).strip() == '':
            break
        headers.append(str(cell_value).strip())
        col_idx += 1
    
    return headers


def clean_value(value):
    """
    Clean and normalize cell value
    
    Args:
        value: Cell value
    
    Returns:
        str or None: Cleaned value
    """
    if value is None:
        return None
    
    if isinstance(value, str):
        value = value.strip()
        if value == '':
            return None
        return value
    
    return value


def extract_image_from_cell(sheet, row, col):
    """
    Extract embedded image from cell and convert to base64
    
    Args:
        sheet: openpyxl worksheet
        row: Row number
        col: Column number (1-indexed)
    
    Returns:
        str or None: Base64 encoded image or None if no image
    """
    # Check if worksheet has images
    if not hasattr(sheet, '_images') or not sheet._images:
        return None
    
    # Get cell coordinate
    cell_coord = f"{get_column_letter(col)}{row}"
    
    # Look for images in this cell
    for image in sheet._images:
        # Check if image is in this cell
        if hasattr(image, 'anchor'):
            anchor = image.anchor
            if hasattr(anchor, '_from'):
                # Check if image starts in this cell
                if anchor._from.col == col - 1 and anchor._from.row == row - 1:
                    # Get image data
                    if hasattr(image, '_data'):
                        image_data = image._data()
                        # Convert to base64
                        return base64.b64encode(image_data).decode('utf-8')
    
    return None


def parse_data_rows(sheet, headers, start_row=11, max_rows=1000):
    """
    Parse data rows from Excel
    
    Args:
        sheet: openpyxl worksheet
        headers: List of header names
        start_row: First data row (default: 11)
        max_rows: Maximum rows to parse (default: 1000)
    
    Returns:
        list: List of dictionaries, one per row
    """
    rows = []
    row_num = start_row
    empty_rows = 0
    max_empty_rows = 5  # Stop after 5 consecutive empty rows
    
    while row_num < start_row + max_rows and empty_rows < max_empty_rows:
        row_data = {}
        is_empty = True
        
        for col_idx, header in enumerate(headers, start=1):
            cell_value = sheet.cell(row=row_num, column=col_idx).value
            cleaned_value = clean_value(cell_value)
            
            if cleaned_value is not None:
                is_empty = False
                row_data[header] = cleaned_value
            
            # Special handling for Image column
            if header == 'Image' and cleaned_value is None:
                image_base64 = extract_image_from_cell(sheet, row_num, col_idx)
                if image_base64:
                    row_data[header] = image_base64
                    is_empty = False
        
        if is_empty:
            empty_rows += 1
        else:
            empty_rows = 0
            row_data['_excel_row'] = row_num  # Track Excel row number for error messages
            rows.append(row_data)
        
        row_num += 1
    
    return rows


def parse_excel_file(file_content, filename=''):
    """
    Parse entire Excel file
    
    Args:
        file_content: Binary file content
        filename: Original filename (for error messages)
    
    Returns:
        dict: {
            'product_type': str,
            'headers': list,
            'rows': list,
            'filename': str,
        }
    
    Raises:
        ValueError: If file cannot be parsed
    """
    try:
        # Load workbook
        file_stream = io.BytesIO(file_content)
        workbook = load_workbook(file_stream, data_only=True)
        sheet = workbook.active
        
        # Detect product type
        product_type = detect_product_type(sheet)
        
        # Parse headers
        headers = parse_headers(sheet, header_row=10)
        
        if not headers:
            raise ValueError("No headers found in row 10")
        
        # Parse data rows
        rows = parse_data_rows(sheet, headers, start_row=11)
        
        return {
            'product_type': product_type,
            'headers': headers,
            'rows': rows,
            'filename': filename,
            'total_rows': len(rows),
        }
    
    except Exception as e:
        raise ValueError(f"Error parsing Excel file '{filename}': {str(e)}")
